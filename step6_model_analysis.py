"""
Step 6 — Walk-Forward Model Analysis
=====================================
Reads step 4 outputs and produces comprehensive analysis of the
4-day walk-forward simulation including equity curve, per-day
performance, cumulative returns, band breakdown, and trade stats.

Reads from: E:\\CVD data\\Models\\production\\
  sim_trades.csv          all test days combined
  sim_trades_day{N}.csv   per-day trades
  equity_curve.csv        day-by-day equity log
  feature_importance.csv
  model_report.txt

Outputs: E:\\CVD data\\Models\\production\\analysis\\
  model_summary.txt
  equity_curve.png          walk-forward equity + daily PnL bars
  daily_performance.png     per-day WR / PF / trade count
  band_analysis.png         per-band across all days
  feature_importance.png    top-20 features
  trade_characteristics.png hold / exit / direction / symbol
  confidence_calibration.png

Usage:
  python step6_model_analysis.py
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.ticker import FuncFormatter
from pathlib import Path
warnings.filterwarnings("ignore")

# =========================================================
# USER SETTINGS
# =========================================================
from config import MODEL_DIR, ANALYSIS_DIR, FEATURES_FILE, GZ_TEST_DIR
INPUT_DIR     = MODEL_DIR
OUTPUT_DIR    = ANALYSIS_DIR
FEATURES_FILE = FEATURES_FILE
N_TEST = len(list(Path(GZ_TEST_DIR).glob("*.gz")))

os.makedirs(OUTPUT_DIR, exist_ok=True)

CONFIDENCE_BANDS = [
    {"name": "Band4_strong",   "min_conf": 0.65, "max_conf": 0.70, "units": 4,
     "sl_pct": 1.0, "tp_pct": 0.30,
     "pressure_thresh_long": 0.80, "pressure_thresh_short": 0.50},
    {"name": "Band3_moderate", "min_conf": 0.60, "max_conf": 0.65, "units": 3,
     "sl_pct": 1.0, "tp_pct": 0.10,
     "pressure_thresh_long": 0.30, "pressure_thresh_short": 0.20},
    {"name": "Band2_base",     "min_conf": 0.575,"max_conf": 0.60, "units": 2,
     "sl_pct": 1.0, "tp_pct": 0.10,
     "pressure_thresh_long": 0.20, "pressure_thresh_short": 0.10},
    {"name": "Band1_minimum",  "min_conf": 0.55, "max_conf": 0.575,"units": 1,
     "sl_pct": 1.0, "tp_pct": 0.10,
     "pressure_thresh_long": 0.00, "pressure_thresh_short": 0.00},
]


# =========================================================
# LOAD INPUTS
# =========================================================
def load_inputs():
    paths = {
        "sim":    os.path.join(INPUT_DIR, "sim_trades.csv"),
        "equity": os.path.join(INPUT_DIR, "equity_curve.csv"),
        "imp":    os.path.join(INPUT_DIR, "feature_importance.csv"),
        "report": os.path.join(INPUT_DIR, "model_report.txt"),
    }
    for k, p in paths.items():
        if not os.path.exists(p):
            raise FileNotFoundError(f"{p} not found. Run step4 first.")

    sim    = pd.read_csv(paths["sim"])
    equity = pd.read_csv(paths["equity"])
    imp    = pd.read_csv(paths["imp"])
    report = {}
    with open(paths["report"], encoding="utf-8") as f:
        for line in f:
            if ":" in line and not line.startswith("=") and not line.startswith("CVD"):
                k, _, v = line.partition(":")
                report[k.strip()] = v.strip()

    # Load per-day files
    day_dfs = {}
    for d in range(1, N_TEST + 1):
        p = os.path.join(INPUT_DIR, f"sim_trades_day{d}.csv")
        if os.path.exists(p):
            day_dfs[d] = pd.read_csv(p)

    print(f"  sim_trades     : {len(sim):,} trades, {len(sim.columns)} cols")
    print(f"  equity_curve   : {len(equity)} days")
    print(f"  day files      : {len(day_dfs)} loaded")
    print(f"  feature_imp    : {len(imp)} features")
    return sim, equity, imp, report, day_dfs


# =========================================================
# METRIC HELPERS
# =========================================================
def calc_metrics(pnls):
    arr = np.array(pnls, dtype=float)
    if len(arr) == 0:
        return {"n":0,"pnl":0,"wr":0,"pf":0,"sh":0,"dd":0,"cal":0}
    gw = arr[arr>0].sum(); gl = np.abs(arr[arr<0]).sum()
    eq = pd.Series(arr.cumsum())
    dd = float((eq - eq.cummax()).min())
    return {
        "n":   len(arr),
        "pnl": float(arr.sum()),
        "wr":  float((arr>0).mean()),
        "pf":  float(gw/(gl+1e-8)),
        "sh":  float(arr.mean()/(arr.std()+1e-8)),
        "dd":  dd,
        "cal": float(arr.sum()/(abs(dd)+1e-8)),
    }

dollar_fmt = FuncFormatter(lambda x, _: f"${x:,.0f}")


# =========================================================
# PEAK SIMULTANEOUS DEPLOYMENT HELPER
# =========================================================
def calc_peak_deployed(sim):
    """
    Reconstructs tick-by-tick deployment per day using entry_idx and bars_held.
    Returns dict: date_tag → peak_deployed_dollar
    """
    import bisect
    results = {}
    for day, grp in sim.groupby("date_tag"):
        grp = grp.copy()
        grp["position_dollar"] = grp["entry_price"] * grp["units"]
        grp["exit_idx"] = grp["entry_idx"] + grp["bars_held"]
        events = []
        for _, row in grp.iterrows():
            events.append((int(row["entry_idx"]), +row["position_dollar"]))
            events.append((int(row["exit_idx"]),  -row["position_dollar"]))
        events.sort(key=lambda x: x[0])
        deployed = 0.0
        peak     = 0.0
        for _, delta in events:
            deployed += delta
            if deployed > peak:
                peak = deployed
        results[day] = peak
    return results


# =========================================================
# CHART 1 — WALK-FORWARD EQUITY CURVE
# =========================================================
def plot_equity_curve(sim, equity_df, report):
    base_equity  = float(report.get("BASE EQUITY",  "5000000").replace(",",""))
    final_equity = float(report.get("FINAL EQUITY", "5000000").replace(",",""))
    net_profit   = float(report.get("NET PROFIT",   "0").replace(",",""))
    net_pct      = float(report.get("NET PCT",      "0").replace("%",""))

    # Compute peak deployed per day
    peak_deployed = calc_peak_deployed(sim)

    fig = plt.figure(figsize=(16, 14))
    gs  = gridspec.GridSpec(4, 1, height_ratios=[3, 1.2, 1, 1], hspace=0.40)
    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1])
    ax3 = fig.add_subplot(gs[2])
    ax4 = fig.add_subplot(gs[3])

    fig.suptitle(
        f"Walk-Forward Equity Curve — {len(equity_df)} Test Days\n"
        f"Start: ${base_equity:,.0f}  →  Final: ${final_equity:,.0f}  "
        f"Net: ${net_profit:+,.2f}  ({net_pct:+.2f}%)",
        fontsize=13, fontweight="bold"
    )

    # Panel 1: cumulative equity line
    eq_points = [base_equity] + list(equity_df["equity_end"].values)
    x_eq      = range(len(eq_points))
    color     = "green" if net_profit >= 0 else "red"
    ax1.plot(x_eq, eq_points, color=color, linewidth=2.5, marker="o",
             markersize=7, label="Equity")
    ax1.axhline(base_equity, color="grey", linewidth=1, linestyle="--",
                label=f"Start ${base_equity:,.0f}")
    ax1.fill_between(x_eq, eq_points, base_equity,
                     where=[e >= base_equity for e in eq_points],
                     alpha=0.15, color="green")
    ax1.fill_between(x_eq, eq_points, base_equity,
                     where=[e < base_equity for e in eq_points],
                     alpha=0.15, color="red")
    for i, (x, eq) in enumerate(zip(x_eq[1:], eq_points[1:]), 1):
        ax1.annotate(f"${eq:,.0f}", (x, eq),
                     textcoords="offset points", xytext=(0, 10),
                     ha="center", fontsize=9, fontweight="bold")
    ax1.set_xticks(list(x_eq))
    ax1.set_xticklabels(["Start"] + list(equity_df["date_tag"]), fontsize=9)
    ax1.yaxis.set_major_formatter(dollar_fmt)
    ax1.set_ylabel("Portfolio Equity ($)")
    ax1.legend(fontsize=9); ax1.grid(True, alpha=0.2)

    # Panel 2: daily PnL bars
    days    = equity_df["date_tag"].values
    day_pnl = equity_df["day_pnl"].values
    colors  = ["green" if p >= 0 else "red" for p in day_pnl]
    bars    = ax2.bar(range(len(days)), day_pnl, color=colors, alpha=0.85, width=0.6)
    ax2.axhline(0, color="black", linewidth=0.8)
    ax2.set_xticks(range(len(days))); ax2.set_xticklabels(days, fontsize=9)
    ax2.yaxis.set_major_formatter(dollar_fmt)
    ax2.set_ylabel("Daily PnL ($)")
    ax2.set_title("Daily PnL", fontsize=10)
    for bar, val in zip(bars, day_pnl):
        ax2.text(bar.get_x() + bar.get_width()/2,
                 val + (max(day_pnl) - min(day_pnl)) * 0.04 * (1 if val >= 0 else -1),
                 f"${val:+,.0f}", ha="center", fontsize=8, fontweight="bold")
    ax2.grid(True, alpha=0.2)

    # Panel 3: equity scalar
    scalars = equity_df["equity_scalar"].values
    bar_c   = ["green" if s >= 1 else "salmon" for s in scalars]
    ax3.bar(range(len(days)), scalars, color=bar_c, alpha=0.8, width=0.6)
    ax3.axhline(1.0, color="black", linewidth=1, linestyle="--", label="Baseline")
    ax3.set_xticks(range(len(days))); ax3.set_xticklabels(days, fontsize=9)
    ax3.set_ylabel("Position Scalar"); ax3.set_title("Position Size Scalar", fontsize=10)
    sc_min = min(scalars); sc_max = max(scalars)
    sc_pad = max((sc_max - sc_min) * 0.4, 0.01)
    ax3.set_ylim(sc_min - sc_pad, sc_max + sc_pad * 1.5)
    ax3.legend(fontsize=8); ax3.grid(True, alpha=0.2)
    for i, (s, d) in enumerate(zip(scalars, days)):
        ax3.text(i, s + sc_pad * 0.15, f"{s:.3f}", ha="center", fontsize=8)

    # Panel 4: peak simultaneous deployment per day
    peak_vals  = [peak_deployed.get(d, 0) for d in days]
    peak_pcts  = [p / (equity_df.loc[equity_df["date_tag"]==d, "equity_start"].values[0])
                  * 100 if len(equity_df.loc[equity_df["date_tag"]==d]) > 0 else 0
                  for d, p in zip(days, peak_vals)]
    overall_peak = max(peak_vals) if peak_vals else 0
    ax4.bar(range(len(days)), peak_vals, color="steelblue", alpha=0.85, width=0.6)
    ax4.set_xticks(range(len(days))); ax4.set_xticklabels(days, fontsize=9)
    ax4.yaxis.set_major_formatter(dollar_fmt)
    ax4.set_ylabel("Peak Deployed ($)")
    ax4.set_title(f"Peak Simultaneous Deployment per Day  "
                  f"(Overall peak: ${overall_peak:,.0f})", fontsize=10)
    ax4.grid(True, alpha=0.2)
    for i, (v, pct) in enumerate(zip(peak_vals, peak_pcts)):
        ax4.text(i, v + overall_peak * 0.02, f"${v:,.0f}\n({pct:.1f}%)",
                 ha="center", fontsize=8, fontweight="bold")

    plt.savefig(os.path.join(OUTPUT_DIR, "equity_curve.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  equity_curve.png")


# =========================================================
# CHART 2 — DAILY PERFORMANCE
# =========================================================
def plot_daily_performance(day_dfs, equity_df):
    if not day_dfs:
        return

    days    = sorted(day_dfs.keys())
    n       = len(days)
    labels  = [equity_df.loc[equity_df["day_num"]==d, "date_tag"].values[0]
               if d in equity_df["day_num"].values else f"Day {d}"
               for d in days]

    wrs, pfs, ns, accs, sharpes = [], [], [], [], []
    for d in days:
        g   = day_dfs[d]
        m   = calc_metrics(g["pnl"].values)
        wrs.append(m["wr"])
        pfs.append(min(m["pf"], 10))   # cap at 10 for display
        ns.append(m["n"])
        sharpes.append(m["sh"])
        row = equity_df[equity_df["day_num"] == d]
        accs.append(float(row["dir_acc"].values[0]) if len(row) else 0)

    x = range(n)
    fig, axes = plt.subplots(2, 3, figsize=(18, 10))
    fig.suptitle("Daily Performance — Walk-Forward Test Days", fontsize=13)

    # Win rate
    cols = ["green" if w >= 0.4 else "tomato" for w in wrs]
    axes[0,0].bar(x, wrs, color=cols, alpha=0.85)
    axes[0,0].axhline(0.5, color="black", linewidth=1, linestyle="--")
    axes[0,0].set_xticks(list(x)); axes[0,0].set_xticklabels(labels, fontsize=9)
    axes[0,0].set_title("Win Rate"); axes[0,0].set_ylabel("Win Rate")
    axes[0,0].yaxis.set_major_formatter(FuncFormatter(lambda v,_: f"{v:.0%}"))
    axes[0,0].grid(True, alpha=0.2)
    for xi, v in zip(x, wrs):
        axes[0,0].text(xi, v+0.005, f"{v:.1%}", ha="center", fontsize=9)

    # Profit factor
    cols = ["green" if p >= 1 else "tomato" for p in pfs]
    axes[0,1].bar(x, pfs, color=cols, alpha=0.85)
    axes[0,1].axhline(1.0, color="black", linewidth=1, linestyle="--")
    axes[0,1].set_xticks(list(x)); axes[0,1].set_xticklabels(labels, fontsize=9)
    axes[0,1].set_title("Profit Factor (capped at 10)")
    axes[0,1].grid(True, alpha=0.2)
    for xi, v in zip(x, pfs):
        axes[0,1].text(xi, v+0.05, f"{v:.2f}", ha="center", fontsize=9)

    # Trade count
    axes[0,2].bar(x, ns, color="steelblue", alpha=0.85)
    axes[0,2].set_xticks(list(x)); axes[0,2].set_xticklabels(labels, fontsize=9)
    axes[0,2].set_title("Trade Count"); axes[0,2].set_ylabel("Trades")
    axes[0,2].grid(True, alpha=0.2)
    for xi, v in zip(x, ns):
        axes[0,2].text(xi, v+20, f"{v:,}", ha="center", fontsize=9)

    # Sharpe
    cols = ["green" if s >= 0 else "tomato" for s in sharpes]
    axes[1,0].bar(x, sharpes, color=cols, alpha=0.85)
    axes[1,0].axhline(0, color="black", linewidth=1)
    axes[1,0].set_xticks(list(x)); axes[1,0].set_xticklabels(labels, fontsize=9)
    axes[1,0].set_title("Sharpe Ratio"); axes[1,0].grid(True, alpha=0.2)
    for xi, v in zip(x, sharpes):
        axes[1,0].text(xi, v+0.001, f"{v:.4f}", ha="center", fontsize=8)

    # Direction accuracy
    axes[1,1].bar(x, accs, color="mediumpurple", alpha=0.85)
    axes[1,1].set_xticks(list(x)); axes[1,1].set_xticklabels(labels, fontsize=9)
    axes[1,1].set_title("Direction Model Accuracy")
    axes[1,1].yaxis.set_major_formatter(FuncFormatter(lambda v,_: f"{v:.0%}"))
    axes[1,1].grid(True, alpha=0.2)
    for xi, v in zip(x, accs):
        axes[1,1].text(xi, v+0.003, f"{v:.1%}", ha="center", fontsize=9)

    # Equity scalar
    scalars = [equity_df.loc[equity_df["day_num"] == d, "equity_scalar"].values[0]
               if d in equity_df["day_num"].values else 1.0
               for d in days]
    cols = ["green" if s >= 1 else "salmon" for s in scalars]
    axes[1,2].bar(x, scalars, color=cols, alpha=0.85)
    axes[1,2].axhline(1.0, color="black", linewidth=1, linestyle="--")
    axes[1,2].set_xticks(list(x)); axes[1,2].set_xticklabels(labels, fontsize=9)
    axes[1,2].set_title("Position Scalar (equity/base)"); axes[1,2].grid(True, alpha=0.2)
    sc_min2 = min(scalars); sc_max2 = max(scalars)
    sc_pad2 = max((sc_max2 - sc_min2) * 0.4, 0.01)
    axes[1,2].set_ylim(sc_min2 - sc_pad2, sc_max2 + sc_pad2 * 1.5)
    for xi, v in zip(x, scalars):
        axes[1,2].text(xi, v + sc_pad2 * 0.15, f"{v:.3f}", ha="center", fontsize=9)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "daily_performance.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  daily_performance.png")


# =========================================================
# CHART 3 — BAND ANALYSIS (all days combined + per day)
# =========================================================
def plot_band_analysis(sim, day_dfs):
    bands_present = [b for b in CONFIDENCE_BANDS
                     if sim[sim["tier"] == b["name"]].shape[0] > 0]
    if not bands_present:
        return

    fig, axes = plt.subplots(2, 3, figsize=(18, 11))
    fig.suptitle("Confidence Band Analysis — All Test Days Combined", fontsize=13)
    x      = range(len(bands_present))
    labels = [f"{b['name']}\n({b['units']}u  TP={b['tp_pct']}%)"
              for b in bands_present]

    wrs, ppts, ns, pnls, pfs = [], [], [], [], []
    for b in bands_present:
        t   = sim[sim["tier"] == b["name"]]
        m   = calc_metrics(t["pnl"].values)
        wrs.append(m["wr"]); ppts.append(m["pnl"]/len(t))
        ns.append(len(t)); pnls.append(m["pnl"]); pfs.append(m["pf"])

    # WR
    axes[0,0].bar(x, wrs, color=["green" if w>=0.4 else "tomato" for w in wrs],
                  alpha=0.85)
    axes[0,0].axhline(0.5, color="black", linewidth=1, linestyle="--")
    axes[0,0].set_xticks(list(x)); axes[0,0].set_xticklabels(labels, fontsize=8)
    axes[0,0].set_title("Win Rate by Band")
    axes[0,0].yaxis.set_major_formatter(FuncFormatter(lambda v,_: f"{v:.0%}"))
    axes[0,0].grid(True, alpha=0.2)
    for xi, v in zip(x, wrs):
        axes[0,0].text(xi, v+0.005, f"{v:.1%}", ha="center", fontsize=8)

    # PF
    axes[0,1].bar(x, [min(p,10) for p in pfs],
                  color=["green" if p>=1 else "tomato" for p in pfs], alpha=0.85)
    axes[0,1].axhline(1, color="black", linewidth=1, linestyle="--")
    axes[0,1].set_xticks(list(x)); axes[0,1].set_xticklabels(labels, fontsize=8)
    axes[0,1].set_title("Profit Factor by Band (capped 10)")
    axes[0,1].grid(True, alpha=0.2)
    for xi, v in zip(x, pfs):
        axes[0,1].text(xi, min(v,10)+0.05, f"{v:.2f}", ha="center", fontsize=8)

    # Total PnL
    axes[0,2].bar(x, pnls, color=["green" if p>=0 else "tomato" for p in pnls],
                  alpha=0.85)
    axes[0,2].axhline(0, color="black", linewidth=1)
    axes[0,2].set_xticks(list(x)); axes[0,2].set_xticklabels(labels, fontsize=8)
    axes[0,2].set_title("Total PnL by Band")
    axes[0,2].yaxis.set_major_formatter(dollar_fmt)
    axes[0,2].grid(True, alpha=0.2)
    for xi, v in zip(x, pnls):
        axes[0,2].text(xi, v + max(pnls)*0.02, f"${v:,.0f}", ha="center", fontsize=8)

    # PnL per trade
    axes[1,0].bar(x, ppts, color=["green" if p>=0 else "tomato" for p in ppts],
                  alpha=0.85)
    axes[1,0].axhline(0, color="black", linewidth=1)
    axes[1,0].set_xticks(list(x)); axes[1,0].set_xticklabels(labels, fontsize=8)
    axes[1,0].set_title("PnL per Trade by Band")
    axes[1,0].grid(True, alpha=0.2)
    for xi, v in zip(x, ppts):
        axes[1,0].text(xi, v + 0.001, f"${v:.4f}", ha="center", fontsize=8)

    # Trade count
    axes[1,1].bar(x, ns, color="steelblue", alpha=0.85)
    axes[1,1].set_xticks(list(x)); axes[1,1].set_xticklabels(labels, fontsize=8)
    axes[1,1].set_title("Trade Count by Band"); axes[1,1].grid(True, alpha=0.2)
    for xi, v in zip(x, ns):
        axes[1,1].text(xi, v+30, f"{v:,}", ha="center", fontsize=8)

    # PnL by band per day (stacked bars)
    if day_dfs:
        day_nums  = sorted(day_dfs.keys())
        bar_width = 0.8 / max(len(day_nums), 1)
        band_colors = ["#2196F3","#4CAF50","#FF9800","#9C27B0","#F44336","#00BCD4"]
        for bi, b in enumerate(bands_present):
            day_pnls = []
            for d in day_nums:
                g = day_dfs[d]
                day_pnls.append(g[g["tier"]==b["name"]]["pnl"].sum())
            xd = [i + bi*bar_width for i in range(len(day_nums))]
            axes[1,2].bar(xd, day_pnls,
                          width=bar_width*0.9,
                          color=band_colors[bi % len(band_colors)],
                          alpha=0.8, label=b["name"].replace("Band","B"))
        axes[1,2].axhline(0, color="black", linewidth=0.8)
        axes[1,2].set_xticks([i + bar_width*len(bands_present)/2
                               for i in range(len(day_nums))])
        axes[1,2].set_xticklabels([f"Day {d}" for d in day_nums], fontsize=9)
        axes[1,2].set_title("PnL by Band per Day")
        axes[1,2].yaxis.set_major_formatter(dollar_fmt)
        axes[1,2].legend(fontsize=7); axes[1,2].grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "band_analysis.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  band_analysis.png")


# =========================================================
# CHART 4 — FEATURE IMPORTANCE
# =========================================================
def plot_feature_importance(imp):
    top = imp.head(20).copy()
    top["pct"] = top["gain"] / imp["gain"].sum() * 100

    fig, ax = plt.subplots(figsize=(12, 8))
    colors  = plt.cm.RdYlGn(np.linspace(0.3, 0.9, len(top)))[::-1]
    bars    = ax.barh(range(len(top)), top["gain"].values,
                      color=colors, alpha=0.85, edgecolor="white")
    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(top["feature"].values, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("XGBoost Gain")
    ax.set_title("Feature Importance — Top 20 by XGBoost Gain", fontsize=11)
    ax.grid(True, alpha=0.2, axis="x")
    for i, (bar, row) in enumerate(zip(bars, top.itertuples())):
        ax.text(bar.get_width() + top["gain"].max()*0.01,
                bar.get_y() + bar.get_height()/2,
                f"{row.gain:.1f}  ({row.pct:.1f}%)",
                va="center", fontsize=8)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "feature_importance.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  feature_importance.png")


# =========================================================
# CHART 5 — TRADE CHARACTERISTICS
# =========================================================
def plot_trade_characteristics(sim):
    fig, axes = plt.subplots(2, 3, figsize=(18, 10))
    fig.suptitle("Trade Characteristics — All Test Days", fontsize=13)

    # Hold time
    axes[0,0].hist(sim["bars_held"], bins=50, color="steelblue", alpha=0.8)
    axes[0,0].axvline(sim["bars_held"].median(), color="red", linewidth=1.5,
                      linestyle="--", label=f"Median={sim['bars_held'].median():.0f}")
    axes[0,0].set_title("Hold Time Distribution (bars)")
    axes[0,0].set_xlabel("Bars held"); axes[0,0].legend(fontsize=9)
    axes[0,0].grid(True, alpha=0.2)

    # Exit reason
    reasons = sim.groupby("exit_reason")["pnl"].agg(["sum","mean","count"])
    cols_r  = ["green" if v >= 0 else "tomato" for v in reasons["sum"]]
    axes[0,1].bar(reasons.index, reasons["sum"], color=cols_r, alpha=0.85)
    axes[0,1].set_title("Total PnL by Exit Reason")
    axes[0,1].yaxis.set_major_formatter(dollar_fmt)
    axes[0,1].grid(True, alpha=0.2)
    for i, (idx, row) in enumerate(reasons.iterrows()):
        axes[0,1].text(i, row["sum"] + abs(reasons["sum"].max())*0.02,
                       f"n={int(row['count']):,}\n${row['mean']:.3f}/t",
                       ha="center", va="bottom", fontsize=7)

    # PnL distribution
    wins   = sim[sim["pnl"] > 0]["pnl"]
    losses = sim[sim["pnl"] < 0]["pnl"]
    axes[0,2].hist(wins,   bins=50, color="green",  alpha=0.6, label=f"Wins n={len(wins):,}")
    axes[0,2].hist(losses, bins=50, color="tomato", alpha=0.6, label=f"Losses n={len(losses):,}")
    axes[0,2].axvline(0, color="black", linewidth=1)
    axes[0,2].set_title("PnL Distribution per Trade")
    axes[0,2].legend(fontsize=9); axes[0,2].grid(True, alpha=0.2)

    # Per-symbol
    if "symbol" in sim.columns:
        sym_pnl = (sim.groupby("symbol")["pnl"].sum()
                   .sort_values(ascending=False).head(15))
        axes[1,0].barh(range(len(sym_pnl)), sym_pnl.values,
                       color=["green" if v>=0 else "tomato" for v in sym_pnl.values],
                       alpha=0.85)
        axes[1,0].set_yticks(range(len(sym_pnl)))
        axes[1,0].set_yticklabels(sym_pnl.index, fontsize=9)
        axes[1,0].invert_yaxis()
        axes[1,0].axvline(0, color="black", linewidth=0.8)
        axes[1,0].set_title("PnL by Symbol (top 15)")
        axes[1,0].xaxis.set_major_formatter(dollar_fmt)
        axes[1,0].grid(True, alpha=0.2)

    # Direction
    dir_stats = sim.groupby("direction").agg(
        n=("pnl","count"), pnl=("pnl","sum"),
        wr=("pnl", lambda x: (x>0).mean())
    )
    dir_names  = {-1:"Short", 1:"Long"}
    dir_labels = [dir_names.get(d, str(d)) for d in dir_stats.index]
    axes[1,1].bar(dir_labels, dir_stats["pnl"],
                  color=["green" if p>=0 else "tomato" for p in dir_stats["pnl"]],
                  alpha=0.85)
    axes[1,1].set_title("Total PnL by Direction")
    axes[1,1].yaxis.set_major_formatter(dollar_fmt)
    axes[1,1].grid(True, alpha=0.2)
    for i, (lbl, row) in enumerate(zip(dir_labels, dir_stats.itertuples())):
        axes[1,1].text(i, row.pnl + abs(dir_stats["pnl"].max())*0.02,
                       f"n={row.n:,}\nWR={row.wr:.1%}",
                       ha="center", va="bottom", fontsize=9)

    # Confidence distribution
    axes[1,2].hist(sim["confidence"], bins=40, color="steelblue", alpha=0.8)
    for b in CONFIDENCE_BANDS:
        axes[1,2].axvline(b["min_conf"], color="red", linewidth=0.8,
                          linestyle="--", alpha=0.6)
        axes[1,2].text(b["min_conf"], axes[1,2].get_ylim()[1]*0.9,
                       f"{b['units']}u", ha="center", fontsize=7, color="red")
    axes[1,2].set_title("Confidence Distribution of Traded Signals")
    axes[1,2].set_xlabel("Confidence"); axes[1,2].grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "trade_characteristics.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  trade_characteristics.png")


# =========================================================
# CHART 6 — CONFIDENCE CALIBRATION
# =========================================================
def plot_confidence_calibration(sim):
    bins = np.arange(0.575, 0.70, 0.01)
    labels, wrs, ns, ppts = [], [], [], []
    for lo, hi in zip(bins[:-1], bins[1:]):
        t = sim[(sim["confidence"] >= lo) & (sim["confidence"] < hi)]
        if len(t) < 5: continue
        labels.append(f"{lo:.3f}")
        wrs.append((t["pnl"] > 0).mean())
        ns.append(len(t))
        ppts.append(t["pnl"].sum() / len(t))

    if not labels:
        return

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 9), sharex=True)
    fig.suptitle("Confidence Calibration — Higher confidence = higher accuracy?",
                 fontsize=12)
    x = list(range(len(labels)))

    # Panel 1: win rate
    cols = ["green" if w >= 0.5 else "tomato" for w in wrs]
    ax1.bar(x, wrs, color=cols, alpha=0.75)
    ax1.axhline(0.5, color="black", linewidth=1, linestyle="--")
    if len(wrs) > 2:
        z = np.polyfit(x, wrs, 1)
        p = np.poly1d(z)
        ax1.plot(x, p(x), "navy", linewidth=2,
                 label=f"Trend (slope={z[0]*1000:.2f} per 0.1)")
    ax1.set_xticks(x); ax1.set_xticklabels(labels, rotation=45, fontsize=8)
    ax1.set_ylabel("Win rate")
    ax1.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:.0%}"))
    ax1.set_title("Win Rate by Confidence Bucket")
    ax1.legend(fontsize=9); ax1.grid(True, alpha=0.2)
    for xi, v in zip(x, wrs):
        ax1.text(xi, v + 0.005, f"{v:.1%}", ha="center", fontsize=8)

    # Panel 2: PnL per trade — independent axis, no twinx distortion
    cols2  = ["green" if p >= 0 else "tomato" for p in ppts]
    bars   = ax2.bar(x, ppts, color=cols2, alpha=0.75)
    ax2.axhline(0, color="black", linewidth=1, linestyle="--")

    # Set y-limits explicitly so small positives don't look negative
    ymax = max(ppts) * 1.35 if max(ppts) > 0 else 0.1
    ymin = min(min(ppts) * 1.35, -0.01)
    ax2.set_ylim(ymin, ymax)

    # Value labels on each bar
    for xi, v in zip(x, ppts):
        offset = ymax * 0.03 if v >= 0 else ymin * 0.03
        ax2.text(xi, v + offset, f"${v:.4f}", ha="center", fontsize=8,
                 color="darkgreen" if v >= 0 else "darkred", fontweight="bold")

    # Trade count as text annotations instead of twinx to avoid axis conflicts
    for xi, n in zip(x, ns):
        ax2.text(xi, ymin + (ymax - ymin) * 0.04,
                 f"n={n:,}", ha="center", fontsize=8, color="black", fontweight="bold")

    ax2.set_xticks(x); ax2.set_xticklabels(labels, rotation=45, fontsize=8)
    ax2.set_ylabel("PnL per trade ($)")
    ax2.set_title("PnL per Trade by Confidence Bucket  (n= shown at bottom)")
    ax2.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "confidence_calibration.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  confidence_calibration.png")


# =========================================================
# CHART 7 — TIME-OF-DAY (SESSION PROGRESS) PERFORMANCE
# =========================================================
def plot_session_progress(sim):
    """
    Uses entry_idx normalised per day as a proxy for session progress.
    Splits into session quintiles: Open, Mid-Morning, Midday, Afternoon, Close.
    """
    df = sim.copy()
    # Normalise entry_idx within each day to [0, 1] session progress
    day_max = df.groupby("date_tag")["entry_idx"].transform("max")
    df["session_pct"] = df["entry_idx"] / (day_max + 1)

    session_labels = ["Open\n0-20%", "Mid-Morn\n20-40%", "Midday\n40-60%",
                      "Afternoon\n60-80%", "Close\n80-100%"]
    session_edges  = [0.0, 0.2, 0.4, 0.6, 0.8, 1.01]

    wrs, ppts, ns, pnls, pfs = [], [], [], [], []
    for lo, hi in zip(session_edges[:-1], session_edges[1:]):
        t = df[(df["session_pct"] >= lo) & (df["session_pct"] < hi)]
        ns.append(len(t))
        wrs.append((t["pnl"] > 0).mean() if len(t) > 0 else 0)
        ppts.append(t["pnl"].mean() if len(t) > 0 else 0)
        pnls.append(t["pnl"].sum())
        gw = t[t["pnl"] > 0]["pnl"].sum(); gl = t[t["pnl"] < 0]["pnl"].abs().sum()
        pfs.append(gw / (gl + 1e-8))

    x = np.arange(len(session_labels))
    fig, axes = plt.subplots(2, 3, figsize=(18, 11))
    fig.suptitle("Session Progress Performance — When During the Day is Edge Strongest?",
                 fontsize=13, fontweight="bold")

    # Win rate
    ax = axes[0, 0]
    cols = ["green" if w >= 0.5 else "tomato" for w in wrs]
    ax.bar(x, wrs, color=cols, alpha=0.85)
    ax.axhline(0.5, color="black", linewidth=1, linestyle="--")
    ax.set_xticks(x); ax.set_xticklabels(session_labels, fontsize=9)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:.0%}"))
    ax.set_title("Win Rate by Session"); ax.grid(True, alpha=0.2)
    for xi, v in zip(x, wrs):
        ax.text(xi, v + 0.005, f"{v:.1%}", ha="center", fontsize=9, fontweight="bold")

    # Profit factor
    ax = axes[0, 1]
    cols = ["green" if p >= 1 else "tomato" for p in pfs]
    ax.bar(x, [min(p, 10) for p in pfs], color=cols, alpha=0.85)
    ax.axhline(1.0, color="black", linewidth=1, linestyle="--")
    ax.set_xticks(x); ax.set_xticklabels(session_labels, fontsize=9)
    ax.set_title("Profit Factor by Session"); ax.grid(True, alpha=0.2)
    for xi, v in zip(x, pfs):
        ax.text(xi, min(v, 10) + 0.03, f"{v:.3f}", ha="center", fontsize=9, fontweight="bold")

    # Total PnL
    ax = axes[0, 2]
    cols = ["green" if p >= 0 else "tomato" for p in pnls]
    ax.bar(x, pnls, color=cols, alpha=0.85)
    ax.axhline(0, color="black", linewidth=1)
    ax.set_xticks(x); ax.set_xticklabels(session_labels, fontsize=9)
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_title("Total PnL by Session"); ax.grid(True, alpha=0.2)
    for xi, v in zip(x, pnls):
        offset = max(abs(p) for p in pnls) * 0.03
        ax.text(xi, v + (offset if v >= 0 else -offset),
                f"${v:+,.0f}", ha="center", fontsize=9, fontweight="bold")

    # PnL per trade
    ax = axes[1, 0]
    cols = ["green" if p >= 0 else "tomato" for p in ppts]
    ax.bar(x, ppts, color=cols, alpha=0.85)
    ax.axhline(0, color="black", linewidth=1)
    ax.set_xticks(x); ax.set_xticklabels(session_labels, fontsize=9)
    ax.set_title("PnL per Trade by Session"); ax.grid(True, alpha=0.2)
    for xi, v in zip(x, ppts):
        ax.text(xi, v + 0.003, f"${v:.4f}", ha="center", fontsize=9)

    # Trade count
    ax = axes[1, 1]
    ax.bar(x, ns, color="steelblue", alpha=0.85)
    ax.set_xticks(x); ax.set_xticklabels(session_labels, fontsize=9)
    ax.set_title("Trade Count by Session"); ax.grid(True, alpha=0.2)
    for xi, v in zip(x, ns):
        ax.text(xi, v + max(ns) * 0.02, f"{v:,}", ha="center", fontsize=9)

    # Per-day session heatmap-style: PnL by session × day
    ax = axes[1, 2]
    days = sorted(df["date_tag"].unique())
    bar_width = 0.8 / max(len(session_labels), 1)
    session_colors = ["#FF9800", "#4CAF50", "#2196F3", "#9C27B0", "#F44336"]
    for si, (lo, hi) in enumerate(zip(session_edges[:-1], session_edges[1:])):
        day_pnls = []
        for d in days:
            g = df[(df["date_tag"] == d) & (df["session_pct"] >= lo) & (df["session_pct"] < hi)]
            day_pnls.append(g["pnl"].sum())
        xd = [i + si * bar_width for i in range(len(days))]
        ax.bar(xd, day_pnls, width=bar_width * 0.9,
               color=session_colors[si], alpha=0.8,
               label=session_labels[si].replace("\n", " "))
    ax.axhline(0, color="black", linewidth=0.8)
    ax.set_xticks([i + bar_width * len(session_labels) / 2 for i in range(len(days))])
    ax.set_xticklabels(days, fontsize=9)
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_title("PnL by Session × Day"); ax.legend(fontsize=7); ax.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "session_progress.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  session_progress.png")


# =========================================================
# CHART 8 — GENERAL TRADE FREQUENCY ANALYSIS
# =========================================================
def plot_tick_frequency_analysis(sim):
    """
    Uses the raw tick count per symbol from clean_features.csv (test days only)
    as the measure of general trade frequency — NOT the model's executed trade count.
    This answers: does the model perform better on stocks with more market activity?
    """
    def pf(pnls):
        arr = np.array(pnls)
        gw = arr[arr > 0].sum()
        gl = np.abs(arr[arr < 0]).sum()
        return gw / (gl + 1e-8)

    # ---- Load general tick counts from clean_features.csv ----
    test_days = sorted(sim["date_tag"].unique())
    general_ticks = None
    try:
        # Only load symbol + date_tag columns for efficiency
        feat_df = pd.read_csv(FEATURES_FILE, usecols=["symbol", "date_tag"])
        feat_test = feat_df[feat_df["date_tag"].isin(test_days)]
        general_ticks = feat_test.groupby("symbol").size().reset_index(name="general_freq")
        print(f"    (loaded general tick counts from clean_features.csv — "
              f"{len(general_ticks)} symbols, {len(feat_test):,} total ticks)")
    except Exception as e:
        print(f"    WARNING: Could not load {FEATURES_FILE}: {e}")
        print(f"    Falling back to model trade count as proxy")

    # ---- Model performance per symbol ----
    sym_stats = sim.groupby("symbol").agg(
        model_trades=("pnl", "count"),
        total_pnl=("pnl", "sum"),
        wr=("pnl", lambda x: (x > 0).mean()),
        pnl_per_trade=("pnl", "mean"),
        avg_bars_held=("bars_held", "mean"),
    ).reset_index()

    sym_stats["pf"] = sym_stats["symbol"].apply(
        lambda s: pf(sim[sim["symbol"] == s]["pnl"].values)
    )

    # ---- Merge general tick counts ----
    if general_ticks is not None:
        sym_stats = sym_stats.merge(general_ticks, on="symbol", how="left")
        sym_stats["general_freq"] = sym_stats["general_freq"].fillna(sym_stats["model_trades"])
    else:
        sym_stats["general_freq"] = sym_stats["model_trades"]

    sym_stats = sym_stats.sort_values("general_freq", ascending=False)
    freq_col = "general_freq"
    freq_label = "General Trade Frequency (ticks in test period)"

    median_freq = sym_stats[freq_col].median()
    high = sym_stats[sym_stats[freq_col] >= median_freq].copy()
    low  = sym_stats[sym_stats[freq_col] <  median_freq].copy()

    high_sim = sim[sim["symbol"].isin(high["symbol"])]
    low_sim  = sim[sim["symbol"].isin(low["symbol"])]
    high_pf  = pf(high_sim["pnl"].values)
    low_pf   = pf(low_sim["pnl"].values)
    high_wr  = (high_sim["pnl"] > 0).mean()
    low_wr   = (low_sim["pnl"] > 0).mean()

    fig, axes = plt.subplots(2, 3, figsize=(18, 12))
    fig.suptitle(
        f"General Trade Frequency Analysis — High vs Low Frequency Symbols\n"
        f"Split at median {median_freq:,.0f} ticks/symbol  |  "
        f"High: {len(high)} symbols  Low: {len(low)} symbols",
        fontsize=13, fontweight="bold"
    )

    # Panel 1: PF by symbol (sorted by general frequency)
    ax = axes[0, 0]
    colors = ["steelblue" if t >= median_freq else "coral"
              for t in sym_stats[freq_col]]
    ax.barh(range(len(sym_stats)), sym_stats["pf"].values,
            color=colors, alpha=0.85)
    ax.set_yticks(range(len(sym_stats)))
    ax.set_yticklabels(
        [f"{s}  ({f:,.0f})" for s, f in
         zip(sym_stats["symbol"], sym_stats[freq_col])],
        fontsize=8)
    ax.invert_yaxis()
    ax.axvline(1.0, color="black", linewidth=1, linestyle="--")
    ax.set_title("Profit Factor by Symbol\n(sorted by general frequency, count shown)")
    ax.grid(True, alpha=0.2)
    from matplotlib.patches import Patch
    ax.legend(handles=[Patch(color="steelblue", label="High freq"),
                       Patch(color="coral",     label="Low freq")], fontsize=8)
    for i, v in enumerate(sym_stats["pf"].values):
        ax.text(v + 0.01, i, f"{v:.3f}", va="center", fontsize=7)

    # Panel 2: Win rate by symbol
    ax = axes[0, 1]
    colors = ["steelblue" if t >= median_freq else "coral"
              for t in sym_stats[freq_col]]
    ax.barh(range(len(sym_stats)), sym_stats["wr"].values, color=colors, alpha=0.85)
    ax.set_yticks(range(len(sym_stats)))
    ax.set_yticklabels(sym_stats["symbol"].values, fontsize=8)
    ax.invert_yaxis()
    ax.axvline(0.5, color="black", linewidth=1, linestyle="--")
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:.0%}"))
    ax.set_title("Win Rate by Symbol")
    ax.grid(True, alpha=0.2)

    # Panel 3: PF by symbol sorted by PF
    ax = axes[0, 2]
    sym_sorted = sym_stats.sort_values("pf", ascending=False)
    colors = ["steelblue" if t >= median_freq else "coral"
              for t in sym_sorted[freq_col]]
    ax.barh(range(len(sym_sorted)), sym_sorted["pf"].values,
            color=colors, alpha=0.85)
    ax.set_yticks(range(len(sym_sorted)))
    ax.set_yticklabels(sym_sorted["symbol"].values, fontsize=8)
    ax.invert_yaxis()
    ax.axvline(1.0, color="black", linewidth=1, linestyle="--")
    ax.set_title("Profit Factor by Symbol (sorted by PF)")
    ax.grid(True, alpha=0.2)
    for i, v in enumerate(sym_sorted["pf"].values):
        ax.text(v + 0.01, i, f"{v:.3f}", va="center", fontsize=7)

    # Panel 4: Summary comparison — high vs low
    ax = axes[1, 0]
    categories = ["Profit Factor", "Win Rate %", "PnL/Trade", "Model Trades"]
    high_vals = [high_pf, high_wr * 100,
                 high_sim["pnl"].sum() / len(high_sim), len(high_sim)]
    low_vals  = [low_pf,  low_wr  * 100,
                 low_sim["pnl"].sum()  / len(low_sim),  len(low_sim)]
    x = np.arange(len(categories))
    w = 0.35
    ax.bar(x - w/2, high_vals, w, label="High freq", color="steelblue", alpha=0.85)
    ax.bar(x + w/2, low_vals,  w, label="Low freq",  color="coral",     alpha=0.85)
    ax.set_xticks(x); ax.set_xticklabels(categories, fontsize=9)
    ax.set_title("High vs Low General Frequency — Summary")
    ax.legend(); ax.grid(True, alpha=0.2)
    for xi, (hv, lv) in enumerate(zip(high_vals, low_vals)):
        ax.text(xi - w/2, hv * 1.02, f"{hv:,.2f}", ha="center", fontsize=7, color="steelblue")
        ax.text(xi + w/2, lv * 1.02, f"{lv:,.2f}", ha="center", fontsize=7, color="coral")

    # Panel 5: General frequency vs PF scatter
    ax = axes[1, 1]
    colors = ["steelblue" if t >= median_freq else "coral"
              for t in sym_stats[freq_col]]
    ax.scatter(sym_stats[freq_col], sym_stats["pf"],
               c=colors, s=80, alpha=0.85, zorder=3)
    ax.axvline(median_freq, color="black", linewidth=1, linestyle="--",
               label=f"Median ({median_freq:,.0f})")
    ax.axhline(1.0, color="black", linewidth=0.8, linestyle="--")
    for _, row in sym_stats.iterrows():
        ax.annotate(row["symbol"], (row[freq_col], row["pf"]),
                    textcoords="offset points", xytext=(4, 2), fontsize=7)
    ax.set_xlabel(freq_label)
    ax.set_ylabel("Profit Factor")
    ax.set_title("General Trade Frequency vs Profit Factor")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.2)

    # Panel 6: Cumulative PnL — high vs low over all trades
    ax = axes[1, 2]
    high_trades = high_sim.sort_values(["date_tag", "entry_idx"]).reset_index(drop=True)
    low_trades  = low_sim.sort_values( ["date_tag", "entry_idx"]).reset_index(drop=True)
    ax.plot(np.arange(len(high_trades)), high_trades["pnl"].cumsum(),
            color="steelblue", linewidth=2,
            label=f"High freq ({len(high)} syms, PF={high_pf:.3f})")
    ax.plot(np.arange(len(low_trades)),  low_trades["pnl"].cumsum(),
            color="coral",     linewidth=2,
            label=f"Low freq ({len(low)} syms, PF={low_pf:.3f})")
    ax.axhline(0, color="black", linewidth=0.8)
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_xlabel("Trade number")
    ax.set_ylabel("Cumulative PnL ($)")
    ax.set_title("Cumulative PnL — High vs Low General Frequency")
    ax.legend(fontsize=9); ax.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "tick_frequency_analysis.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  tick_frequency_analysis.png")


# =========================================================
# CHART 9 — INTRADAY CUMULATIVE PnL (one line per day)
# =========================================================
def plot_intraday_cumulative(sim, day_dfs):
    if not day_dfs:
        return

    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    fig.suptitle("Intraday Cumulative PnL — Trade-by-Trade Within Each Day",
                 fontsize=13, fontweight="bold")

    day_colors = plt.cm.tab10(np.linspace(0, 1, max(len(day_dfs), 1)))

    # Panel 1: absolute cumulative PnL
    ax = axes[0]
    for i, (d, g) in enumerate(sorted(day_dfs.items())):
        g_sorted = g.sort_values("entry_idx").reset_index(drop=True)
        cum_pnl  = g_sorted["pnl"].cumsum()
        label    = g_sorted["date_tag"].iloc[0] if "date_tag" in g_sorted.columns else f"Day {d}"
        ax.plot(range(len(cum_pnl)), cum_pnl, linewidth=1.8,
                color=day_colors[i], label=label, alpha=0.85)
    ax.axhline(0, color="black", linewidth=0.8, linestyle="--")
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_xlabel("Trade number (within day)")
    ax.set_ylabel("Cumulative PnL ($)")
    ax.set_title("Cumulative PnL per Day")
    ax.legend(fontsize=9); ax.grid(True, alpha=0.2)

    # Panel 2: normalised (per-day drawdown from peak)
    ax = axes[1]
    for i, (d, g) in enumerate(sorted(day_dfs.items())):
        g_sorted = g.sort_values("entry_idx").reset_index(drop=True)
        cum_pnl  = g_sorted["pnl"].cumsum()
        dd       = cum_pnl - cum_pnl.cummax()
        label    = g_sorted["date_tag"].iloc[0] if "date_tag" in g_sorted.columns else f"Day {d}"
        ax.fill_between(range(len(dd)), dd, 0, alpha=0.25, color=day_colors[i])
        ax.plot(range(len(dd)), dd, linewidth=1.2, color=day_colors[i], label=label)
    ax.axhline(0, color="black", linewidth=0.8)
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_xlabel("Trade number (within day)")
    ax.set_ylabel("Drawdown from Day Peak ($)")
    ax.set_title("Intraday Drawdown per Day")
    ax.legend(fontsize=9); ax.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "intraday_cumulative.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  intraday_cumulative.png")


# =========================================================
# CHART 10 — ROLLING METRICS (stability of edge)
# =========================================================
def plot_rolling_metrics(sim):
    if len(sim) < 400:
        return

    WINDOW = 200
    ordered = sim.sort_values(["date_tag", "entry_idx"]).reset_index(drop=True)
    pnl     = ordered["pnl"].values

    # Rolling arrays
    r_wr  = pd.Series(pnl > 0).rolling(WINDOW).mean().values
    r_ppt = pd.Series(pnl).rolling(WINDOW).mean().values

    # Rolling PF
    gw_roll = pd.Series(np.where(pnl > 0, pnl, 0)).rolling(WINDOW).sum().values
    gl_roll = pd.Series(np.where(pnl < 0, np.abs(pnl), 0)).rolling(WINDOW).sum().values
    r_pf    = gw_roll / (gl_roll + 1e-8)

    x = np.arange(len(pnl))

    fig, axes = plt.subplots(3, 1, figsize=(16, 12), sharex=True)
    fig.suptitle(f"Rolling {WINDOW}-Trade Metrics — Is the Edge Stable?",
                 fontsize=13, fontweight="bold")

    # Mark day boundaries
    day_boundaries = []
    prev_day = None
    for i, row in ordered.iterrows():
        if row["date_tag"] != prev_day:
            if prev_day is not None:
                day_boundaries.append(i)
            prev_day = row["date_tag"]

    # Panel 1: Rolling win rate
    ax = axes[0]
    ax.plot(x, r_wr, color="steelblue", linewidth=1.5)
    ax.axhline(0.5, color="black", linewidth=1, linestyle="--", label="50%")
    ax.fill_between(x, r_wr, 0.5, where=r_wr >= 0.5, alpha=0.15, color="green")
    ax.fill_between(x, r_wr, 0.5, where=r_wr <  0.5, alpha=0.15, color="red")
    for b in day_boundaries:
        ax.axvline(b, color="gray", linewidth=0.5, linestyle=":", alpha=0.5)
    ax.set_ylabel("Win Rate")
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:.0%}"))
    ax.set_title(f"Rolling {WINDOW}-Trade Win Rate")
    ax.legend(fontsize=9); ax.grid(True, alpha=0.2)

    # Panel 2: Rolling PF
    ax = axes[1]
    ax.plot(x, r_pf, color="green", linewidth=1.5)
    ax.axhline(1.0, color="black", linewidth=1, linestyle="--", label="Breakeven")
    ax.fill_between(x, r_pf, 1.0, where=r_pf >= 1.0, alpha=0.15, color="green")
    ax.fill_between(x, r_pf, 1.0, where=r_pf <  1.0, alpha=0.15, color="red")
    for b in day_boundaries:
        ax.axvline(b, color="gray", linewidth=0.5, linestyle=":", alpha=0.5)
    ax.set_ylabel("Profit Factor")
    ax.set_ylim(0, min(r_pf[WINDOW:].max() * 1.3, 5) if len(r_pf) > WINDOW else 5)
    ax.set_title(f"Rolling {WINDOW}-Trade Profit Factor")
    ax.legend(fontsize=9); ax.grid(True, alpha=0.2)

    # Panel 3: Rolling PnL per trade
    ax = axes[2]
    ax.plot(x, r_ppt, color="mediumpurple", linewidth=1.5)
    ax.axhline(0, color="black", linewidth=1, linestyle="--")
    ax.fill_between(x, r_ppt, 0, where=r_ppt >= 0, alpha=0.15, color="green")
    ax.fill_between(x, r_ppt, 0, where=r_ppt <  0, alpha=0.15, color="red")
    for b in day_boundaries:
        ax.axvline(b, color="gray", linewidth=0.5, linestyle=":", alpha=0.5)
    ax.set_ylabel("PnL per Trade ($)")
    ax.set_xlabel("Trade number (chronological)")
    ax.set_title(f"Rolling {WINDOW}-Trade PnL per Trade")
    ax.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "rolling_metrics.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  rolling_metrics.png")


# =========================================================
# CHART 11 — UNDERWATER / DRAWDOWN PLOT
# =========================================================
def plot_underwater(sim):
    ordered = sim.sort_values(["date_tag", "entry_idx"]).reset_index(drop=True)
    pnl     = ordered["pnl"].values
    cum_pnl = pd.Series(pnl).cumsum()
    dd      = cum_pnl - cum_pnl.cummax()
    max_dd  = dd.min()
    max_dd_idx = dd.idxmin()

    fig, axes = plt.subplots(2, 1, figsize=(16, 10),
                             gridspec_kw={"height_ratios": [2, 1.5]})
    fig.suptitle(f"Equity & Drawdown — {len(sim):,} Trades  |  "
                 f"Max Drawdown: ${max_dd:,.2f}",
                 fontsize=13, fontweight="bold")

    # Panel 1: Cumulative PnL
    ax = axes[0]
    ax.plot(cum_pnl, color="green", linewidth=1.5)
    ax.fill_between(cum_pnl.index, cum_pnl, 0,
                    where=cum_pnl >= 0, alpha=0.15, color="green")
    ax.fill_between(cum_pnl.index, cum_pnl, 0,
                    where=cum_pnl < 0, alpha=0.15, color="red")
    ax.axhline(0, color="black", linewidth=0.8, linestyle="--")
    # Mark day boundaries
    prev_day = None
    for i, row in ordered.iterrows():
        if row["date_tag"] != prev_day:
            if prev_day is not None:
                ax.axvline(i, color="gray", linewidth=0.5, linestyle=":", alpha=0.5)
            prev_day = row["date_tag"]
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_ylabel("Cumulative PnL ($)")
    ax.set_title("Cumulative PnL (all test days sequential)")
    ax.grid(True, alpha=0.2)

    # Panel 2: Underwater
    ax = axes[1]
    ax.fill_between(dd.index, dd, 0, color="tomato", alpha=0.5)
    ax.plot(dd, color="darkred", linewidth=1.0)
    ax.axhline(0, color="black", linewidth=0.8)
    ax.axhline(max_dd, color="red", linewidth=1, linestyle="--",
               label=f"Max DD: ${max_dd:,.2f} (trade #{max_dd_idx:,})")
    # Mark day boundaries
    prev_day = None
    for i, row in ordered.iterrows():
        if row["date_tag"] != prev_day:
            if prev_day is not None:
                ax.axvline(i, color="gray", linewidth=0.5, linestyle=":", alpha=0.5)
            prev_day = row["date_tag"]
    ax.yaxis.set_major_formatter(dollar_fmt)
    ax.set_xlabel("Trade number (chronological)")
    ax.set_ylabel("Drawdown ($)")
    ax.set_title("Underwater Plot — Distance from Equity Peak")
    ax.legend(fontsize=10); ax.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "underwater.png"),
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  underwater.png")


# =========================================================
# TEXT SUMMARY
# =========================================================
def write_summary(sim, equity_df, imp, report, day_dfs):
    SEP  = "=" * 62
    SEP2 = "-" * 62
    path = os.path.join(OUTPUT_DIR, "model_summary.txt")

    base_equity  = float(report.get("BASE EQUITY",  "5000000").replace(",",""))
    final_equity = float(report.get("FINAL EQUITY", "5000000").replace(",",""))
    net_profit   = float(report.get("NET PROFIT",   "0").replace(",",""))
    net_pct      = float(report.get("NET PCT",      "0").replace("%",""))

    m = calc_metrics(sim["pnl"].values)

    with open(path, "w", encoding="utf-8") as f:
        f.write(f"{SEP}\nCVD WALK-FORWARD MODEL SUMMARY\n{SEP}\n\n")

        f.write(f"EQUITY SUMMARY\n{SEP2}\n")
        f.write(f"  Starting equity : ${base_equity:>12,.2f}\n")
        f.write(f"  Final equity    : ${final_equity:>12,.2f}\n")
        f.write(f"  Net profit      : ${net_profit:>+12,.2f}  ({net_pct:+.2f}%)\n\n")

        f.write(f"DAY-BY-DAY RESULTS\n{SEP2}\n")
        f.write(f"  {'Day':>3}  {'Date':12}  {'Start':>12}  {'Scalar':>6}  "
                f"{'n':>6}  {'PnL':>10}  {'%':>7}  {'End':>12}  "
                f"{'WR':>6}  {'PF':>7}  {'Sh':>8}  {'Acc':>5}\n")
        f.write(f"  {'-'*100}\n")
        for _, r in equity_df.iterrows():
            d = int(r.day_num)
            if d in day_dfs:
                g  = day_dfs[d]
                dm = calc_metrics(g["pnl"].values)
                wr_s = f"{dm['wr']:.1%}"; pf_s = f"{dm['pf']:.3f}"; sh_s = f"{dm['sh']:.4f}"
            else:
                wr_s = "—"; pf_s = "—"; sh_s = "—"
            f.write(f"  {d:>3}  {r.date_tag:12}  "
                    f"${r.equity_start:>11,.2f}  {r.equity_scalar:>6.3f}  "
                    f"{int(r.n_trades):>6,}  ${r.day_pnl:>+9,.2f}  "
                    f"{r.day_pct:>+6.2f}%  ${r.equity_end:>11,.2f}  "
                    f"{wr_s:>6}  {pf_s:>7}  {sh_s:>8}  {r.dir_acc:>5.1%}\n")
        f.write("\n")

        f.write(f"COMBINED PERFORMANCE (all test days)\n{SEP2}\n")
        f.write(f"  Total trades   : {m['n']:>10,}\n")
        f.write(f"  Total PnL      : ${m['pnl']:>12,.2f}\n")
        f.write(f"  Win rate       : {m['wr']:>12.1%}\n")
        f.write(f"  Profit factor  : {m['pf']:>12.3f}\n")
        f.write(f"  Sharpe         : {m['sh']:>12.4f}\n")
        f.write(f"  Max drawdown   : ${m['dd']:>12,.2f}\n")
        n_test_days = len(equity_df)
        f.write(f"  Calmar         : {m['cal']:>12.3f}  ({n_test_days} test days — not annualised)\n\n")

        f.write(f"CONFIDENCE BANDS\n{SEP2}\n")
        f.write(f"  {'Band':22}  {'SL':>4}  {'TP':>5}  "
                f"{'n':>7}  {'WR':>6}  {'PF':>7}  "
                f"{'PnL':>12}  {'PnL/t':>9}  {'hold':>6}\n")
        f.write(f"  {'-'*90}\n")
        for b in CONFIDENCE_BANDS:
            t = sim[sim["tier"] == b["name"]]
            if len(t) == 0:
                continue
            else:
                bm = calc_metrics(t["pnl"].values)
                f.write(f"  {b['name']:22}  {b['sl_pct']:>3g}%  {b['tp_pct']:>4g}%  "
                        f"{bm['n']:>7,}  {bm['wr']:>6.1%}  {bm['pf']:>7.3f}  "
                        f"${bm['pnl']:>11,.2f}  ${bm['pnl']/bm['n']:>8.4f}  "
                        f"{t['bars_held'].mean():>5.1f}b\n")
        f.write("\n")

        f.write(f"EXIT REASONS\n{SEP2}\n")
        for reason, g in sim.groupby("exit_reason"):
            em = calc_metrics(g["pnl"].values)
            f.write(f"  {reason:16}: n={em['n']:>6,}  WR={em['wr']:.1%}  "
                    f"PF={em['pf']:.3f}  PnL=${em['pnl']:,.2f}  "
                    f"avg=${em['pnl']/em['n']:.4f}\n")
        f.write("\n")

        f.write(f"LONG vs SHORT\n{SEP2}\n")
        for d, label in [(1,"Long"),(-1,"Short")]:
            g  = sim[sim["direction"] == d]
            dm = calc_metrics(g["pnl"].values)
            f.write(f"  {label}: n={dm['n']:>6,}  WR={dm['wr']:.1%}  "
                    f"PF={dm['pf']:.3f}  PnL=${dm['pnl']:,.2f}\n")
        f.write("\n")

        if "symbol" in sim.columns:
            f.write(f"PER-SYMBOL\n{SEP2}\n")
            f.write(f"  {'Symbol':12}  {'n':>6}  {'WR':>6}  {'PF':>7}  "
                    f"{'PnL':>12}  {'PnL/t':>9}\n")
            f.write(f"  {'-'*60}\n")
            sym_stats = sim.groupby("symbol").agg(
                n=("pnl","count"), wr=("pnl", lambda x:(x>0).mean()),
                pnl=("pnl","sum")
            )
            sym_stats["pnl_per_trade"] = sym_stats["pnl"] / sym_stats["n"]
            sym_stats = sym_stats.sort_values("pnl_per_trade", ascending=False)
            for sym, row in sym_stats.iterrows():
                bm2 = calc_metrics(sim[sim["symbol"]==sym]["pnl"].values)
                f.write(f"  {sym:12}  {int(row.n):>6,}  {row.wr:>6.1%}  "
                        f"{bm2['pf']:>7.3f}  ${row.pnl:>11,.2f}  "
                        f"${row.pnl/row.n:>8.4f}\n")
            f.write("\n")

        f.write(f"TOP 15 FEATURES\n{SEP2}\n")
        total_gain = imp["gain"].sum()
        for rank, (_, row) in enumerate(imp.head(15).iterrows(), 1):
            pct = row["gain"] / total_gain * 100
            f.write(f"  {rank:>2}. {row['feature']:38}  "
                    f"gain={row['gain']:>9.3f}  ({pct:.1f}%)\n")

    print(f"  model_summary.txt")


# =========================================================
# EXCEL EXPORT
# =========================================================
def export_colour_graded_excel(sim):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed — skipping Excel export")
        return

    BAND_COLOURS = {
        "Band1_minimum":  "FFFFFF",
        "Band2_base":     "FFFF00",
        "Band3_moderate": "00CC44",
        "Band4_strong":   "00FFFF",
    }
    WIN_COL  = "ABFBBE"
    LOSS_COL = "FFA7A7"

    wb = Workbook(); ws = wb.active; ws.title = "Sim Trades"
    headers     = list(sim.columns)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 28

    pnl_col  = headers.index("pnl")  + 1
    tier_col = headers.index("tier") + 1
    conf_col = headers.index("confidence") + 1

    for ri, (_, row) in enumerate(sim.iterrows(), 2):
        is_win   = row["pnl"] > 0
        base_hex = WIN_COL if is_win else LOSS_COL
        conf_hex = BAND_COLOURS.get(str(row.get("tier","")), "FFFFFF")
        for ci, col in enumerate(headers, 1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.font   = Font(size=9)
            cell.alignment = Alignment(
                horizontal="right" if isinstance(val,(int,float)) else "left",
                vertical="center")
            if ci in (conf_col, tier_col):
                cell.fill = PatternFill("solid", fgColor=conf_hex)
            else:
                cell.fill = PatternFill("solid", fgColor=base_hex)
            if col == "pnl":
                cell.number_format = "#,##0.0000"
                cell.font = Font(size=9, bold=True,
                                 color="005500" if is_win else "880000")
            elif col == "confidence":
                cell.number_format = "0.0000"
            elif col in ("entry_price","exit_price"):
                cell.number_format = "#,##0.0000"

    ws.freeze_panes = "A2"
    out_path = os.path.join(OUTPUT_DIR, "sim_trades_coloured.xlsx")
    wb.save(out_path)
    print(f"  sim_trades_coloured.xlsx  ({len(sim):,} rows)")


# =========================================================
# MAIN
# =========================================================
def run():
    print("Loading step 4 outputs...")
    sim, equity_df, imp, report, day_dfs = load_inputs()

    print("\nGenerating charts...")
    plot_equity_curve(sim, equity_df, report)
    plot_daily_performance(day_dfs, equity_df)
    plot_band_analysis(sim, day_dfs)
    plot_feature_importance(imp)
    plot_trade_characteristics(sim)
    plot_confidence_calibration(sim)
    plot_session_progress(sim)
    plot_tick_frequency_analysis(sim)
    plot_intraday_cumulative(sim, day_dfs)
    plot_rolling_metrics(sim)
    plot_underwater(sim)

    print("\nWriting summary...")
    write_summary(sim, equity_df, imp, report, day_dfs)

    print("\nExporting Excel...")
    export_colour_graded_excel(sim)

    print(f"\n{'='*50}")
    print(f"Step 5 complete → {OUTPUT_DIR}")


if __name__ == "__main__":
    run()